import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# 中国各省面积和人口数据（2024年统计数据）
province_data = [
    {"name": "北京市", "area": 16410.54, "population": 2154.2},
    {"name": "天津市", "area": 11966.45, "population": 1560.0},
    {"name": "河北省", "area": 188800.00, "population": 7591.97},
    {"name": "山西省", "area": 156700.00, "population": 3729.22},
    {"name": "内蒙古自治区", "area": 1183000.00, "population": 2539.6},
    {"name": "辽宁省", "area": 148000.00, "population": 4359.3},
    {"name": "吉林省", "area": 187400.00, "population": 2690.73},
    {"name": "黑龙江省", "area": 473000.00, "population": 3751.3},
    {"name": "上海市", "area": 6340.50, "population": 2428.14},
    {"name": "江苏省", "area": 107200.00, "population": 8070.0},
    {"name": "浙江省", "area": 105500.00, "population": 5850.0},
    {"name": "安徽省", "area": 140100.00, "population": 6365.9},
    {"name": "福建省", "area": 124000.00, "population": 3973.0},
    {"name": "江西省", "area": 166900.00, "population": 4666.1},
    {"name": "山东省", "area": 157900.00, "population": 10070.21},
    {"name": "河南省", "area": 167000.00, "population": 9640.0},
    {"name": "湖北省", "area": 185900.00, "population": 5927.0},
    {"name": "湖南省", "area": 211800.00, "population": 6918.38},
    {"name": "广东省", "area": 179700.00, "population": 11521.0},
    {"name": "广西壮族自治区", "area": 237600.00, "population": 4960.0},
    {"name": "海南省", "area": 35400.00, "population": 944.72},
    {"name": "重庆市", "area": 82400.00, "population": 3124.32},
    {"name": "四川省", "area": 486000.00, "population": 8375.0},
    {"name": "贵州省", "area": 176100.00, "population": 3622.95},
    {"name": "云南省", "area": 394100.00, "population": 4858.3},
    {"name": "西藏自治区", "area": 1228400.00, "population": 350.56},
    {"name": "陕西省", "area": 205600.00, "population": 3876.21},
    {"name": "甘肃省", "area": 425800.00, "population": 2647.43},
    {"name": "青海省", "area": 722300.00, "population": 607.82},
    {"name": "宁夏回族自治区", "area": 66400.00, "population": 720.27},
    {"name": "新疆维吾尔自治区", "area": 1664900.00, "population": 2523.22},
    {"name": "香港特别行政区", "area": 1106.66, "population": 750.0},
    {"name": "澳门特别行政区", "area": 32.90, "population": 68.32},
    {"name": "台湾省", "area": 36193.00, "population": 2360.0}
]

def create_excel_file(data, filename="province_area_population.xlsx"):
    """
    创建Excel文件并写入省份数据
    """
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "中国各省面积与人口"
    
    # 设置表头
    headers = ["省份", "面积（平方公里）", "人口（万人）"]
    ws.append(headers)
    
    # 设置表头样式
    header_font = Font(bold=True, color="FFFFFFFF", size=12)
    header_fill = PatternFill(start_color="FF4F81BD", end_color="FF4F81BD", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), 
                         top=Side(style="thin"), bottom=Side(style="thin"))
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # 写入数据
    for row in data:
        ws.append([row["name"], row["area"], row["population"]])
    
    # 设置数据样式
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.alignment = center_align
            cell.border = thin_border
    
    # 设置列宽
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15
    
    # 保存文件
    wb.save(filename)
    print(f"数据已成功写入Excel文件: {filename}")

if __name__ == "__main__":
    create_excel_file(province_data)
